from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
# 从demo.xlsx取一列 复制到 demo-test.xlsx的某一列
wb = load_workbook ('./data/Gemm_demo.xlsx')
ws = wb['Gemm_Generality_prof_hpa_gemm_s']
sheet = wb.active
a = sheet.max_row
b=[]
i=0
while i < a:
    i = i+1
    # print(ws.cell (i,7).value) #这里 cell (i,7) 表示取demo.xlsx第7列
    b.insert(i, ws.cell (i,7).value)

wbtest = load_workbook ('./data/gemm_demo_test.xlsx')
ws = wbtest['Sheet1']
sheet = wb.active
i = 1 # 这里i=0从第0行开始赋值 、i=1从第2行开始复制
while i < a:
    ws.cell (i+1,6).value = b[i] # 这里 cell (i,1) 表示赋值到demo-test.xlsx的第7列
    i = i+1
wbtest.save ('./data/gemm_demo_test.xlsx')