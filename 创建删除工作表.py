import openpyxl
wb = openpyxl.Workbook()
wb.create_sheet(title="销售记录")
wb.create_sheet(index=0,title="养殖技术")
'''# 删除Sheet的工作表
del wb['Sheet']'''
# 写操作
sheet = wb['养殖技术']
cell = sheet['A1']
cell.value = 'hello'
wb.save('./data/第一个工作簿.xlsx')