import openpyxl
# 打开一个存在的工作簿对象
wb = openpyxl.load_workbook('./data/demo.xlsx')
sheet = wb['Sheet1']
'''
q=sheet.columns 
print(list(q))
'''
# 打印第一列的数据
# for cell in list(sheet.columns)[0]:
#     print(cell.value)
print(sheet.max_row,sheet.max_column)