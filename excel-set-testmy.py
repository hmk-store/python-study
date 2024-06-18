import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, Color
#  对于out.xlsx中存在的列 in.xlsx中的列数据均会对齐覆盖
# 第一个Excel文件路径
source_excel_path = './data/in-2.xlsx'
# 第二个Excel文件路径
target_excel_path = './data/out.xlsx'
pd.options.io.excel.xlsx.writer = 'openpyxl'
# 读取第一个Excel文件
source_df = pd.read_excel(source_excel_path)
 
# 读取第二个Excel文件
target_df = pd.read_excel(target_excel_path, engine='openpyxl')
 
# 找到两个DataFrame中都存在的列名
common_columns = source_df.columns.intersection(target_df.columns)
 
# 仅对共同存在的列进行操作
for column in common_columns:
    target_df[column] = source_df[column]
 
# 将修改后的目标DataFrame保存回Excel文件
# 使用openpyxl作为引擎来处理.xlsx文件
target_df.to_excel(target_excel_path, index=False, engine='openpyxl')
 
# 加载工作簿
wb = load_workbook(target_excel_path, data_only=True)  # 使用data_only=True来加快读取速度
 
# 选择工作表，此处假设我们要修改第一个工作表
ws = wb.active
 
# 设置字体样式和对齐方式
for row in ws.iter_rows(min_row=2):  # 从第二行开始，假设第一行是列名
    for cell in row:
        cell.font = Font(name='Calibri', size=11, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center")
 
# 加边框
border_style = Border(
    left=Side(style='thin', color=Color('000000')),
    right=Side(style='thin', color=Color('000000')),
    top=Side(style='thin', color=Color('000000')),
    bottom=Side(style='thin', color=Color('000000'))
)
 
# 获取数据区域的行数和列数
max_row = ws.max_row
max_column = ws.max_column + 10  # openpyxl的列索引是从1开始的
 
# 应用边框样式到所有单元格（除了列名）
for row in ws.iter_rows(min_row=2, max_row=max_row):
    for cell in row:
        cell.border = border_style
 
# 保存工作簿
wb.save(target_excel_path)