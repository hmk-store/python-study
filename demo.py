import openpyxl
# print(openpyxl.__version__)
# 打开一个存在的工作簿对象
wb = openpyxl.load_workbook('./data/demo.xlsx')
# print(wb)
# # 输出： <openpyxl.workbook.workbook.Workbook object at 0x000001E062B8FCB0>
# # 获取工作簿的sheet表的名称
# print(wb.sheetnames)
# # 输出：['Sheet1', 'Sheet2', 'Sheet3']
# # 获取活动表
# print(wb.active)
# # 获取指定工作簿
# sheet = wb['Sheet2']
# print(sheet)
# 获取单元格——确定位置 
'''
方式一：
sheet = wb['Sheet1']
cell =  sheet['A4']
print(cell)
print(cell.value) #值
print(cell.coordinate) #坐标
print(cell.column) #列索引
print(cell.row) #行索引
方式二：
sheet = wb['Sheet1']
cell = sheet.cell(row=6,column=1)
print(cell)
print(cell.value)'''
'''# 利用循环批量获取单元格中的数据
sheet = wb['Sheet1']
for cell_row in sheet['A2':'B4']:
    for cell in cell_row:
        print(cell.coordinate,cell.value)'''

sheet = wb['Sheet1']
# # 输出单元格所有的列
# q=sheet.columns 
# # print(next(q))
print(list(q))
for cell in list(sheet.columns)[2]:
    print(cell.value)