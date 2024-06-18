from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
# 从demo.xlsx取一列 复制到 demo-test.xlsx的某一列
wb = load_workbook ('./data/demo.xlsx')
ws = wb['Sheet1']
sheet = wb.active
a = sheet.max_row
b=[]
i=0
while i < a:
    i = i+1    
    print(ws.cell (i,1).value) # # openpyxl的列索引是从1开始的, 这里 cell (i,1) 表示取demo.xlsx第一列
    b.insert(i, ws.cell (i,1).value)

wbtest = load_workbook ('./data/demo-test.xlsx')
ws = wbtest['Sheet1']
sheet = wb.active
i = 0
while i < a:
    ws.cell (i+1,1).value = b[i] # 这里 cell (i,1) 表示赋值到demo-test.xlsx的第一列
    i = i+1
wbtest.save ('./data/demo-test.xlsx')